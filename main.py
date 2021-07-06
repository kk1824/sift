from glob import glob
import csv
import math
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side

def read_csv(path):
    csv_file = open(path, "r", encoding="utf-8", errors="", newline="" )
    f = csv.reader(csv_file)
    data_list = []      #CSVから読み込んだデータをリストに格納していく
    request_list = []   #希望日時を管理する変数
    remarks_list = []   #備考（欠勤理由）を管理する変数

    #CSVファイルの一行ごと順にlistに格納
    for row in f:
        data_list.append(row)

    year = data_list[0][0].replace("年","")     #年を格納
    month = data_list[0][1].replace("月","")    #月を格納
    day_list = []       #曜日を格納するlist

    #曜日を上から順に格納
    for row_num in range(6,37):
        day_list.append(data_list[row_num][1])

    name_list = []      #名前を格納する変数
    #スプレットシート3行目の名前をlistに格納
    for name in data_list[2]:
        if name != "名前" and name != "":
            name_list.append(name)          #name_listに名前を追加
            request_list.append([name])     #誰のデータなのかわかりやすいように希望日時のlistの先頭にも追加
            remarks_list.append([name])     #誰のデータなのかわかりやすいように備考のlistの先頭にも追加
    col_num = 2
    for num in range(len(name_list)):
        for row_num in range(6,37):
            request_list[num].append(data_list[row_num][col_num])       #希望日時（出社時刻）をlistに追加
            request_list[num].append(data_list[row_num][col_num+1])     #希望日時（退社時刻）をlistに追加
            remarks_list[num].append(data_list[row_num][col_num+2])     #備考をlistに追加
        col_num += 3
    #年，月，曜日のlist，希望日時のlist，備考のlistを返す
    return year,month,day_list,request_list,remarks_list

#セル指定を行うために列のA~ZZまでのlistを作成
def create_col_list():
    col_list = []
    for i in range(65,91):
        col_list.append(chr(i))
    for i in range(65,91):
        for j in range(65,91):
            col_list.append(chr(i)+chr(j))
    return col_list

#エクセルファイル作成
def create_excel(year,month,day_list,request_list,remarks_list):
    file_name = "{0}月シフト申請表.xlsx".format(month)  #ファイルの名前の変数
    wb = openpyxl.Workbook()
    col_list = create_col_list()

    #上旬，下旬，備考の3つのシートを作成
    wb.create_sheet("{0}.{1}月 上旬シフト申請".format(year,month))
    wb.create_sheet("{0}.{1}月 下旬シフト申請".format(year,month))
    wb.create_sheet("{0}.{1}月 欠勤理由".format(year,month))

    #1枚目と2枚目のシートはほぼ同じ処理
    sheet_name_list = ["{0}.{1}月 上旬シフト申請".format(year,month),"{0}.{1}月 下旬シフト申請".format(year,month)]
    for sheet_name in sheet_name_list:
        ws = wb[sheet_name]             #シート指定
        ws.sheet_view.zoomScale = 60    #表示する倍率の指定
        ws.freeze_panes = "B3"          #氏名，日，曜日が常に表示されるように固定

        ws.row_dimensions[1].height = 18.5          #日にちを表示するセルの高さを指定
        ws.row_dimensions[2].height = 15.75         #曜日を表示するセルの高さを指定
        ws.column_dimensions["A"].width = 15.6      #氏名を表示する幅を指定

        #「氏名」書き込み
        ws.merge_cells("A1:A2")
        ws["A1"].value = "氏名"
        ws["A1"].font = Font(size=12, bold=True)    #フォントサイズの変更とボールド
        ws["A1"].alignment = Alignment(horizontal = "center", vertical = "center")  #上下左右中央揃え

        #幅と高さの調整
        row_num = 3
        #名前や希望日時を表示するセルの高さを指定
        for i in range(len(request_list)):
            ws.row_dimensions[row_num].height = 28.5
            row_num += 1
        col_num = 1
        #希望日時を表示するセルの幅を指定
        for i in range(32):
            ws.column_dimensions[col_list[col_num]].width = 6
            col_num += 1

        #日付と曜日を出力するセルの結合
        col_num = 1
        for i in range(16):
            marge_cells = "{0}1:{1}1".format(col_list[col_num],col_list[col_num+1])
            ws.merge_cells(marge_cells)
            marge_cells = "{0}2:{1}2".format(col_list[col_num],col_list[col_num+1])
            ws.merge_cells(marge_cells)
            col_num += 2

        #氏名の一覧出力
        row_num = 3
        for i in range(len(request_list)):
            cell = "A{0}".format(row_num)
            ws[cell].value = request_list[i][0]     #名前の表示
            ws[cell].font = Font(size=12)           #フォントサイズの変更
            row_num += 1

        
        #日にちと曜日の出力
        if sheet_name == "{0}.{1}月 上旬シフト申請".format(year,month):
            date = 0    #上旬は1日から
            max = 31    #15日分
        else:
            date = 15   #下旬は16日から
            max = 33    #31日までの月があるため16日分
        for col_num in range(1,max,2):
            if day_list[date] != "":
                #日にちを表示
                cell = "{0}1".format(col_list[col_num])
                ws[cell].value = date + 1
                ws[cell].font = Font(size=12, bold=True)
                ws[cell].alignment = Alignment(horizontal = "center")
                #曜日を表示
                cell = "{0}2".format(col_list[col_num])
                ws[cell].value = day_list[date]
                #土曜日は青，日曜日は赤で表示
                if day_list[date] == "土":
                    ws[cell].font = Font(color="0000ff", size=12)
                elif day_list[date] == "日":
                    ws[cell].font = Font(color="ff0000", size=12)
                else:
                    ws[cell].font = Font(size=12)
                ws[cell].alignment = Alignment(horizontal = "center")   #左右中央揃え
            date += 1

        #申請時間の出力
        if sheet_name == "{0}.{1}月 上旬シフト申請".format(year,month):
            max = 31    #上旬は15日分
            step = 0
        else:
            max = 33    #下旬は16日分
            step = 30   #下旬は希望日時のlistの後半部分を使う
        list_num = 0
        row_num = 3
        for request in request_list:
            for col_num in range(1,max):
                cell = "{0}{1}".format(col_list[col_num],row_num)
                if request[col_num+step] != "":
                    hour,min = request[col_num+step].split(":")     #時間を「:」で時間と分に分ける
                    #表示の関係上0時を24時に，1時を25時に直す
                    if hour == "0":
                        hour = "24"
                    elif hour == "1":
                        hour = "25"
                    #30分は「.5」で表記
                    if min == "30":
                        min = ".5"
                        hour += min
                    ws[cell].value = float(hour)
                #希望日時が入力されている日に備考欄が入力されている場合，時間を赤で表示する
                if request[col_num+step] != "" and remarks_list[list_num][math.ceil((col_num+step)/2)] != "":
                    ws[cell].font = Font(color="ff0000", size=12)
                else:
                    ws[cell].font = Font(size=12)
                ws[cell].alignment = Alignment(horizontal = "center")
            list_num += 1
            row_num += 1

        #罫線の出力
        border = Border(top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000'), 
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'))
        for row_num in range(1,3):
            for col_num in range(1,34):
                ws.cell(row=row_num ,column=col_num).border = border
        for row_num in range(1,len(request_list)+3):
            ws.cell(row=row_num ,column=1).border = border
        
        border = Border(top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000'), 
                    left=Side(style='thin', color='000000'))
        for row_num in range(3,len(request_list)+3,2):
            for col_num in range(2,34,2):
                ws.cell(row=row_num ,column=col_num).border = border
        border = Border(top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'))
        for row_num in range(3,len(request_list)+3):
            for col_num in range(3,34,2):
                ws.cell(row=row_num ,column=col_num).border = border

    #3枚目のシート
    ws = wb["{0}.{1}月 欠勤理由".format(year,month)]
    ws.sheet_view.zoomScale = 60
    ws.freeze_panes = "C3"

    ws.row_dimensions[1].height = 28.5
    ws.row_dimensions[2].height = 28.5
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 40

    ws.merge_cells("A1:C1")
    ws["A1"].value = "欠勤理由リスト"
    ws["A1"].font = Font(size=12, bold=True)
    ws["A1"].alignment = Alignment(horizontal = "center", vertical = "center")
    ws.merge_cells("A2:B2")
    ws["A2"].value = "日付＼氏名"
    ws["A2"].font = Font(size=12)
    ws["A2"].alignment = Alignment(horizontal = "center", vertical = "center")

    #幅と高さの調整
    row_num = 3
    for i in range(31):
        ws.row_dimensions[row_num].height = 28.5
        row_num += 1
    col_num = 3
    for i in range(len(remarks_list)):
        ws.column_dimensions[col_list[col_num]].width = 40
        col_num += 1

    #日にちと曜日の出力
    for row_num in range(3,34):
            if day_list[row_num - 3] != "":
                cell = "A{0}".format(row_num)
                ws[cell].value = row_num - 2
                ws[cell].font = Font(size=12, bold=True)
                ws[cell].alignment = Alignment(horizontal = "center", vertical = "center")
                cell = "B{0}".format(row_num)
                ws[cell].value = day_list[row_num - 3]
                if day_list[row_num - 3] == "土":
                    ws[cell].font = Font(color="0000ff", size=12)
                elif day_list[row_num - 3] == "日":
                    ws[cell].font = Font(color="ff0000", size=12)
                else:
                    ws[cell].font = Font(size=12)
                ws[cell].alignment = Alignment(horizontal = "center", vertical = "center")
    
    #氏名の一覧出力
    col_num = 2
    for i in range(len(remarks_list)):
        cell = "{0}2".format(col_list[col_num])
        ws[cell].value = remarks_list[i][0]
        ws[cell].font = Font(size=12)
        ws[cell].alignment = Alignment(horizontal = "center", vertical = "center")
        col_num += 1

    list_num = 0
    col_num = 2
    for remarks in remarks_list:
        for row_num in range(3,34):
            cell = "{0}{1}".format(col_list[col_num],row_num)
            ws[cell].value = remarks[row_num-2]
            if remarks[row_num-2] != "" and request_list[list_num][(row_num-2)*2] != "":
                ws[cell].font = Font(color="ff0000", size=12)
            else:
                ws[cell].font = Font(size=12)
        list_num += 1
        col_num += 1

    #罫線の出力
    border = Border(top=Side(style='thin', color='000000'), 
                bottom=Side(style='thin', color='000000'), 
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'))
    for row_num in range(2,34):
        for col_num in range(1,len(request_list)+3):
            ws.cell(row=row_num ,column=col_num).border = border

    #Excelファイルを作成した際に自動生成されるシートを削除
    del wb['Sheet']

    #ファイルを保存して終了
    wb.save(file_name)


if __name__ == "__main__":
    path = glob("*.csv")[0]     #フォルダ内のcsvファイルを取得
    year,month,day_list,request_list,remarks_list = read_csv(path)
    create_excel(year,month,day_list,request_list,remarks_list)
    